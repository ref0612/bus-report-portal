#!/usr/bin/env python3
"""
Excel report generator using openpyxl.
Reads JSON from stdin, writes xlsx to stdout (base64).
"""
import sys, json, base64, io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers as xl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.chart.axis import NumericAxis
from openpyxl.chart.data_source import NumDataSource, NumRef
import logging

# Configure logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# ─── Palette ──────────────────────────────────────────────────────────────────
BD = '1F4E79'; BM = '2E75B6'; BL = 'D5E8F0'
RD = 'C00000'; RL = 'FFE7E7'
AD = 'BF8F00'; AL = 'FFF2CC'
GD = '375623'; GL = 'E2EFDA'
GR = 'F2F2F2'; WH = 'FFFFFF'

# ─── Translations ─────────────────────────────────────────────────────────────
TRANSLATIONS = {
  'en': {
    'sheet1': 'Executive Summary',
    'sheet2': 'Daily Detail',
    'sheet3': 'Gateway Analysis',
    'sheet4': 'Channel & Device',
    'sheet_pend': 'Pend-Failures (URGENT)',
    'title_main': lambda op: f'PAYMENT CONVERSION & FAILURE REPORT — {op.upper()}',
    'title_period': lambda period, today, days: f'Period: {period}  |  Generated: {today}  |  {days} days',
    'kpi_section': 'KEY PERFORMANCE INDICATORS',
    'kpi_hdrs': ['Metric', 'Total', 'Daily Avg.', 'Min/Day', 'Max/Day', 'Peak Day', 'Rate', 'Status'],
    'kpi_sales': 'Completed sales', 'kpi_fail': 'Gateway failures',
    'kpi_pend': 'Pending-failures', 'kpi_aban': 'Abandonments',
    'kpi_cancel': 'Cancellations (excluded)',
    'status_normal': 'NORMAL', 'status_critical': 'CRITICAL',
    'status_warning': 'WARNING', 'status_high': 'HIGH',
    'lost_section': 'ESTIMATED LOST REVENUE (CLP)',
    'lost_hdrs': ['Category', 'Transactions', 'Total Amount (CLP)', '% of Total', 'Avg. Ticket (CLP)', '', '', ''],
    'lost_fail': 'Gateway failures', 'lost_pend': 'Pending-failures',
    'lost_aban': 'Abandonments', 'lost_total': 'TOTAL',
    'daily_title': lambda op: f'DAILY BREAKDOWN — {op.upper()}',
    'daily_hdrs': ['Date','Sales','Failures','Pend.','Abandonments','Not Converted','Fail Rate','Abnd Rate','Note'],
    'obs_crit': '!! Critical', 'obs_high': 'High fails',
    'obs_rate4': 'Rate >4%', 'obs_rate3': 'Rate >3%', 'obs_aban': '!! High abandons',
    'avg_total': 'DAILY AVERAGE', 'na': 'N/A',
    'chart1_title': 'Daily Sales vs Gateway Failures',
    'chart1_y': 'Count', 'chart1_x': 'Date',
    'chart2_title': 'Daily Failure Rate (%)', 'chart2_y': 'Rate %',
    'gw_title': 'PAYMENT GATEWAY ANALYSIS',
    'gw_hdrs': ['Gateway','Direct Failures','Pend. Failures','Abandonments','Total Incidents','% of Total','Diagnosis'],
    'gw_chart': 'Total Incidents by Gateway',
    'diag_dominant': '!! Dominant failure source — urgent action required.',
    'diag_pend_many': 'Multiple pending-failures — review webhooks immediately.',
    'diag_pend_one': 'Pending-failure detected — verify bank reconciliation.',
    'diag_aban': 'High abandonment volume. Review UX/error messaging.',
    'diag_monitor': 'Monitor. No direct failures.',
    'ch_title': 'CHANNEL AND DEVICE ANALYSIS',
    'ch_by_channel': 'FAILURES & ABANDONMENTS BY CHANNEL',
    'ch_hdrs': ['Channel','Failures','% Failures','Abandonments','% Abandons','Note'],
    'ch_note_top': '!! Review SDK integration', 'ch_note_mon': 'Monitor',
    'pl_title': 'PLATFORM DISTRIBUTION',
    'pl_hdrs': ['Platform','Failures','% Failures','Abandonments','% Abandons',''],
    'raw_fail_title': lambda n: f'GATEWAY FAILURES — {n} records',
    'raw_aban_title': lambda n: f'ABANDONMENTS — {n} records',
    'raw_cols': ['PB Code','Issue Date','Origin','Destination','Seat','Operator','Channel','Gateway','PG Status','Platform','Ticket Price (CLP)'],
    'pend_title': 'PENDING-FAILURES (URGENT) — PAYMENT PROCESSED WITHOUT CONFIRMATION',
    'pend_warn': '⚠ URGENT: Customers below may have been charged without receiving their ticket — verify reconciliation immediately.',
    'pend_cols': ['PB Code','Issue Date','Origin','Destination','Seat','Operator','Channel','Gateway','Ticket Price (CLP)','Customer Email'],
  },
  'es': {
    'sheet1': 'Resumen Ejecutivo',
    'sheet2': 'Detalle Diario',
    'sheet3': 'Análisis Gateways',
    'sheet4': 'Canal y Dispositivo',
    'sheet_pend': 'Pend-Fallidos (URGENTE)',
    'title_main': lambda op: f'CONVERSIÓN DE PAGOS E INFORME DE FALLAS — {op.upper()}',
    'title_period': lambda period, today, days: f'Período: {period}  |  Generado: {today}  |  {days} días',
    'kpi_section': 'INDICADORES CLAVE DE DESEMPEÑO',
    'kpi_hdrs': ['Métrica', 'Total', 'Prom. Diario', 'Mín/Día', 'Máx/Día', 'Día Pico', 'Tasa', 'Estado'],
    'kpi_sales': 'Ventas completadas', 'kpi_fail': 'Fallas de gateway',
    'kpi_pend': 'Pendientes-fallidos', 'kpi_aban': 'Abandonos',
    'kpi_cancel': 'Cancelaciones (excluidas)',
    'status_normal': 'NORMAL', 'status_critical': 'CRÍTICO',
    'status_warning': 'ADVERTENCIA', 'status_high': 'ALTO',
    'lost_section': 'INGRESOS ESTIMADOS PERDIDOS (CLP)',
    'lost_hdrs': ['Categoría', 'Transacciones', 'Monto Total (CLP)', '% del Total', 'Ticket Prom. (CLP)', '', '', ''],
    'lost_fail': 'Fallas de gateway', 'lost_pend': 'Pendientes-fallidos',
    'lost_aban': 'Abandonos', 'lost_total': 'TOTAL',
    'daily_title': lambda op: f'DETALLE DIARIO — {op.upper()}',
    'daily_hdrs': ['Fecha','Ventas','Fallas','Pend.','Abandonos','No Convertidos','Tasa Falla','Tasa Aban.','Nota'],
    'obs_crit': '!! Crítico', 'obs_high': 'Fallas altas',
    'obs_rate4': 'Tasa >4%', 'obs_rate3': 'Tasa >3%', 'obs_aban': '!! Abandonos altos',
    'avg_total': 'PROMEDIO DIARIO', 'na': 'N/D',
    'chart1_title': 'Ventas Diarias vs Fallas de Gateway',
    'chart1_y': 'Cantidad', 'chart1_x': 'Fecha',
    'chart2_title': 'Tasa Diaria de Falla (%)', 'chart2_y': 'Tasa %',
    'gw_title': 'ANÁLISIS DE GATEWAYS DE PAGO',
    'gw_hdrs': ['Gateway','Fallas Directas','Pend. Fallidos','Abandonos','Total Incidentes','% del Total','Diagnóstico'],
    'gw_chart': 'Total de Incidentes por Gateway',
    'diag_dominant': '!! Fuente dominante de fallas — acción urgente requerida.',
    'diag_pend_many': 'Múltiples pendientes-fallidos — revisar webhooks de inmediato.',
    'diag_pend_one': 'Pendiente-fallido detectado — verificar reconciliación bancaria.',
    'diag_aban': 'Alto volumen de abandonos. Revisar UX/mensajes de error.',
    'diag_monitor': 'Monitorear. Sin fallas directas.',
    'ch_title': 'ANÁLISIS DE CANAL Y DISPOSITIVO',
    'ch_by_channel': 'FALLAS Y ABANDONOS POR CANAL',
    'ch_hdrs': ['Canal','Fallas','% Fallas','Abandonos','% Abandonos','Nota'],
    'ch_note_top': '!! Revisar integración SDK', 'ch_note_mon': 'Monitorear',
    'pl_title': 'DISTRIBUCIÓN POR PLATAFORMA',
    'pl_hdrs': ['Plataforma','Fallas','% Fallas','Abandonos','% Abandonos',''],
    'raw_fail_title': lambda n: f'FALLAS DE GATEWAY — {n} registros',
    'raw_aban_title': lambda n: f'ABANDONOS — {n} registros',
    'raw_cols': ['Código PB','Fecha Emisión','Origen','Destino','Asiento','Operador','Canal','Gateway','Estado PG','Plataforma','Precio Pasaje (CLP)'],
    'pend_title': 'PENDIENTES-FALLIDOS (URGENTE) — PAGO PROCESADO SIN CONFIRMACIÓN',
    'pend_warn': '⚠ URGENTE: Los clientes a continuación pueden haber sido cobrados sin recibir su pasaje — verificar reconciliación de inmediato.',
    'pend_cols': ['Código PB','Fecha Emisión','Origen','Destino','Asiento','Operador','Canal','Gateway','Precio Pasaje (CLP)','Email Cliente'],
  },
}

def get_t(data):
    lang = data.get('lang', 'en')
    return TRANSLATIONS.get(lang, TRANSLATIONS['en'])



# ─── Style helpers ─────────────────────────────────────────────────────────────
def bdr_side(color='CCCCCC'):
    return Side(style='thin', color=color)

def thin_border(color='CCCCCC'):
    s = bdr_side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def font(name='Arial', size=10, bold=False, color='000000', italic=False):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_cell(ws, row, col, value=None, bg=WH, fg='000000', bold=False,
             sz=10, halign='center', wrap=False, num_fmt=None, border_color='CCCCCC',
             italic=False):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.font      = font(bold=bold, color=fg, size=sz, italic=italic)
    c.fill      = fill(bg)
    c.alignment = align(h=halign, wrap=wrap)
    c.border    = thin_border(border_color)
    if num_fmt:
        c.number_format = num_fmt
    return c

def hdr(ws, row, col, value, bg=BM, fg='FFFFFF', sz=10, bold=True, wrap=False, width=None):
    c = set_cell(ws, row, col, value, bg=bg, fg=fg, bold=bold, sz=sz, wrap=wrap)
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def status_style(status):
    return {
        'CRITICAL': (RD, RL), 'CRÍTICO': (RD, RL),
        'WARNING': (AD, AL),  'ADVERTENCIA': (AD, AL),
        'HIGH':     (AD, AL), 'ALTO': (AD, AL),
        'NORMAL':  (GD, GL),
    }.get(status, ('000000', WH))

# ─── Sheet 1: Executive Summary ───────────────────────────────────────────────

def _bar_labels(fmt=None):
    """Data labels on top of each bar."""
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.position = 'outEnd'
    if fmt:
        dl.numFmt = fmt
    return dl

def _line_labels(fmt=None):
    """Data labels above each marker."""
    dl = DataLabelList()
    dl.showVal = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.position = 't'
    if fmt:
        dl.numFmt = fmt
    return dl

def build_sheet1(wb, data):
    logging.debug(f"Data received in build_sheet1: {data}")
    ws = wb.active
    T = get_t(data)
    ws.title = T['sheet1']
    ws.sheet_view.showGridLines = False

    operator    = data['operator']
    period      = f"{data['periodStart']} – {data['periodEnd']}"
    today       = data['today']
    total_days  = data['totalDays']
    has_sales   = data['hasSales']
    total_sales = data.get('totalSales') or 0
    avg_sales   = data.get('avgSales') or 0
    t_fail      = data['totalFailures']
    t_pend      = data['totalPending']
    t_aban      = data['totalAbandon']
    t_cancel    = data.get('totalCancelled', 0)
    avg_fr      = data['avgFailRate']
    avg_ar      = data['avgAbanRate']
    avg_fail    = data['avgFailures']
    avg_aban    = data['avgAbandon']
    p_fail      = data['priceFailures']
    p_pend      = data['pricePending']
    p_aban      = data['priceAbandon']
    total_lost  = data['totalLost']
    daily       = data['daily']
    peak_fail   = data.get('peakFailDay') or {}
    peak_aban   = data.get('peakAbanDay') or {}

    sales_vals  = [d['sales'] for d in daily] if has_sales else [0]
    fail_vals   = [d['failures'] for d in daily]
    aban_vals   = [d['abandonments'] for d in daily]

    # Title
    ws.merge_cells('A1:H1')
    set_cell(ws, 1, 1, T['title_main'](operator),
             bg=BD, fg='FFFFFF', bold=True, sz=14, halign='center')
    ws.row_dimensions[1].height = 36

    ws.merge_cells('A2:H2')
    set_cell(ws, 2, 1, T['title_period'](period, today, total_days),
             bg=BL, fg='444444', sz=10, halign='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    # KPI Section header
    ws.merge_cells('A4:H4')
    set_cell(ws, 4, 1, T['kpi_section'], bg=BD, fg='FFFFFF', bold=True, sz=11, halign='left')
    ws.row_dimensions[4].height = 22

    kpi_hdrs = T['kpi_hdrs']
    for ci, h in enumerate(kpi_hdrs, 1):
        hdr(ws, 5, ci, h, bg=BM)
    ws.row_dimensions[5].height = 20

    kpi_rows = []
    if has_sales:
        kpi_rows.append((T['kpi_sales'], total_sales, round(avg_sales,1),
                         min(sales_vals), max(sales_vals), '—', '—', 'NORMAL'))
    kpi_rows.append((T['kpi_fail'], t_fail, round(avg_fail,1),
                     min(fail_vals), peak_fail.get('failures',0),
                     peak_fail.get('dateStr','—'), f"{avg_fr:.2f}%", 'CRITICAL'))
    kpi_rows.append((T['kpi_pend'], t_pend, round(t_pend/total_days,1),
                     0, max(d['pending'] for d in daily),
                     '—', f"{avg_fr:.2f}%", 'WARNING'))
    kpi_rows.append((T['kpi_aban'], t_aban, round(avg_aban,1),
                     min(aban_vals), peak_aban.get('abandonments',0),
                     peak_aban.get('dateStr','—'), f"{avg_ar:.2f}%", T['status_high']))
    if t_cancel > 0:
        kpi_rows.append((T['kpi_cancel'], t_cancel, '—', '—', '—', '—', '—', 'INFO'))

    for ri, row in enumerate(kpi_rows, 6):
        row_bg = GR if ri % 2 == 0 else WH
        for ci, val in enumerate(row, 1):
            if ci == 8:  # Status
                fg_c, bg_c = status_style(val)
                set_cell(ws, ri, ci, val, bg=bg_c, fg=fg_c, bold=True, sz=10)
            else:
                set_cell(ws, ri, ci, val, bg=row_bg,
                         halign='left' if ci == 1 else 'center',
                         num_fmt='#,##0' if isinstance(val, int) and ci > 1 else None)
        ws.row_dimensions[ri].height = 18

    # Spacer
    next_r = 6 + len(kpi_rows) + 1
    ws.row_dimensions[next_r - 1].height = 10

    # Lost revenue section
    ws.merge_cells(f'A{next_r}:H{next_r}')
    set_cell(ws, next_r, 1, T['lost_section'], bg=RD, fg='FFFFFF', bold=True, sz=11, halign='left')
    ws.row_dimensions[next_r].height = 22

    lost_hdrs = T['lost_hdrs']
    for ci, h in enumerate(lost_hdrs, 1):
        if ci <= 5:
            hdr(ws, next_r+1, ci, h, bg=RD)
    ws.row_dimensions[next_r+1].height = 20

    lost_data = [
        (T['lost_fail'],  t_fail, p_fail, p_fail/total_lost if total_lost else 0, round(p_fail/t_fail) if t_fail else 0),
        (T['lost_pend'],  t_pend, p_pend, p_pend/total_lost if total_lost else 0, round(p_pend/t_pend) if t_pend else 0),
        (T['lost_aban'],      t_aban, p_aban, p_aban/total_lost if total_lost else 0, round(p_aban/t_aban) if t_aban else 0),
        (T['lost_total'], t_fail+t_pend+t_aban, total_lost, 1.0, round(total_lost/(t_fail+t_pend+t_aban)) if (t_fail+t_pend+t_aban) else 0),
    ]
    for ri2, row in enumerate(lost_data, next_r+2):
        is_total = row[0] == 'TOTAL'
        bg = RL if is_total else (GR if ri2 % 2 == 0 else WH)
        fg = RD if is_total else '000000'
        for ci, val in enumerate(row, 1):
            if ci > 5: break
            fmt = None
            if ci in [3, 5] and isinstance(val, (int, float)): fmt = '#,##0'
            if ci == 4: fmt = '0.0%'
            set_cell(ws, ri2, ci, val, bg=bg, fg=fg, bold=is_total,
                     halign='left' if ci == 1 else 'center', num_fmt=fmt)
        ws.row_dimensions[ri2].height = 18

    # Column widths
    for ci, w in enumerate([28,14,14,12,12,18,10,12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ─── Sheet 2: Daily Detail with charts ────────────────────────────────────────
def build_sheet2(wb, data):
    T = get_t(data)
    ws = wb.create_sheet(T['sheet2'])
    ws.sheet_view.showGridLines = False

    daily     = data['daily']
    has_sales = data['hasSales']
    operator  = data['operator']
    period    = f"{data['periodStart']} – {data['periodEnd']}"
    total_days= data['totalDays']

    ws.merge_cells('A1:I1')
    set_cell(ws, 1, 1, T['daily_title'](operator),
             bg=BD, fg='FFFFFF', bold=True, sz=13, halign='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:I2')
    set_cell(ws, 2, 1, T['title_period'](period, '', total_days).split('  |  Generated')[0] + f'  |  {total_days} ' + ('días' if data.get('lang')=='es' else 'days'),
             bg=BL, fg='555555', sz=9, halign='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    hdrs = T['daily_hdrs']
    for ci, h in enumerate(hdrs, 1):
        hdr(ws, 4, ci, h, bg=BM, wrap=True)
    ws.row_dimensions[4].height = 28

    for ri, d in enumerate(daily, 5):
        is_crit = d['failures'] >= 20 or d['failureRate'] >= 4
        is_warn = d['failures'] >= 10 or d['failureRate'] >= 3
        bg = RL if is_crit else (AL if is_warn else (GR if ri % 2 == 0 else WH))
        obs = []
        if d['failures'] >= 20: obs.append(T['obs_crit'])
        elif d['failures'] >= 10: obs.append(T['obs_high'])
        if d['failureRate'] >= 4: obs.append(T['obs_rate4'])
        elif d['failureRate'] >= 3: obs.append(T['obs_rate3'])
        if d['abandonments'] >= 250: obs.append(T['obs_aban'])

        row_vals = [
            d['dateStr'],
            d['sales'] if has_sales else T['na'],
            d['failures'],
            d['pending'],
            d['abandonments'],
            d['totalNotConverted'],
            d['failureRate'] / 100,
            d['abandonRate'] / 100,
            ' | '.join(obs) or '—',
        ]
        for ci, val in enumerate(row_vals, 1):
            fmt = None
            if ci in [2,3,4,5,6] and isinstance(val, (int, float)): fmt = '#,##0'
            if ci in [7,8]: fmt = '0.00%'
            fg = RD if (is_crit and ci in [3,7]) else '000000'
            set_cell(ws, ri, ci, val, bg=bg, fg=fg, bold=(is_crit and ci in [1,3]),
                     halign='left' if ci in [1,9] else 'center', num_fmt=fmt)
        ws.row_dimensions[ri].height = 18

    # Totals row
    tr = len(daily) + 5
    total_row = [T['avg_total'],
                 f'=AVERAGE(B5:B{tr-1})' if has_sales else T['na'],
                 f'=AVERAGE(C5:C{tr-1})', f'=AVERAGE(D5:D{tr-1})',
                 f'=AVERAGE(E5:E{tr-1})', f'=AVERAGE(F5:F{tr-1})',
                 f'=AVERAGE(G5:G{tr-1})', f'=AVERAGE(H5:H{tr-1})', '']
    fmt_map = {1:None,2:'#,##0',3:'#,##0',4:'#,##0',5:'#,##0',6:'#,##0',7:'0.00%',8:'0.00%',9:None}
    for ci, val in enumerate(total_row, 1):
        set_cell(ws, tr, ci, val, bg=BD, fg='FFFFFF', bold=True,
                 halign='left' if ci==1 else 'center', num_fmt=fmt_map.get(ci))
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = 'A5'

    # Column widths
    for ci, w in enumerate([14,10,12,10,14,16,12,12,26], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    chart_start_row = tr + 2
    cats = Reference(ws, min_col=1, min_row=5, max_row=tr-1)
    is_es = data.get('lang') == 'es'

    # ── CHART 1: Ventas diarias (solo si hay datos de ventas) ─────────────
    if has_sales:
        c_sales = BarChart()
        c_sales.type = 'col'
        c_sales.grouping = 'clustered'
        c_sales.title = 'Ventas Diarias' if is_es else 'Daily Sales'
        c_sales.style = 2
        c_sales.height = 10; c_sales.width = 24
        c_sales.y_axis.title = 'Ventas' if is_es else 'Sales'
        c_sales.y_axis.numFmt = '#,##0'
        c_sales.x_axis.tickLblSkip = max(1, len(daily)//15)
        c_sales.x_axis.tickMarkSkip = max(1, len(daily)//15)
        c_sales.legend = None
        d_sales = Reference(ws, min_col=2, min_row=4, max_row=tr-1)
        c_sales.add_data(d_sales, titles_from_data=True)
        c_sales.set_categories(cats)
        c_sales.series[0].graphicalProperties.solidFill = '2E75B6'
        c_sales.series[0].graphicalProperties.line.solidFill = '1F5490'
        c_sales.series[0].graphicalProperties.line.width = 6000
        c_sales.series[0].dLbls = _bar_labels('#,##0')
        ws.add_chart(c_sales, f'A{chart_start_row}')
        chart_start_row += 14

    # ── CHART 2: Fallas diarias (barras rojas, eje propio) ────────────────
    c_fail = BarChart()
    c_fail.type = 'col'
    c_fail.grouping = 'clustered'
    c_fail.title = 'Fallas de Pasarela por Día' if is_es else 'Daily Gateway Failures'
    c_fail.style = 2
    c_fail.height = 10; c_fail.width = 24
    c_fail.y_axis.title = 'Fallas' if is_es else 'Failures'
    c_fail.y_axis.numFmt = '#,##0'
    c_fail.y_axis.crossAx = 100
    c_fail.x_axis.tickLblSkip = max(1, len(daily)//15)
    c_fail.x_axis.tickMarkSkip = max(1, len(daily)//15)
    c_fail.legend = None
    d_fail = Reference(ws, min_col=3, min_row=4, max_row=tr-1)
    c_fail.add_data(d_fail, titles_from_data=True)
    c_fail.set_categories(cats)
    c_fail.series[0].graphicalProperties.solidFill = 'C00000'
    c_fail.series[0].graphicalProperties.line.solidFill = '900000'
    c_fail.series[0].graphicalProperties.line.width = 6000
    c_fail.series[0].dLbls = _bar_labels('#,##0')
    ws.add_chart(c_fail, f'A{chart_start_row}')
    chart_start_row += 14

    # ── CHART 3: Tasa de falla % (línea) ─────────────────────────────────
    c_rate = LineChart()
    c_rate.title = T['chart2_title']
    c_rate.style = 2
    c_rate.height = 10; c_rate.width = 24
    c_rate.y_axis.title = T['chart2_y']
    c_rate.y_axis.numFmt = '0.00%'
    c_rate.y_axis.crossAx = 100
    c_rate.x_axis.tickLblSkip = max(1, len(daily)//15)
    c_rate.x_axis.tickMarkSkip = max(1, len(daily)//15)
    c_rate.legend = None
    d_rate = Reference(ws, min_col=7, min_row=4, max_row=tr-1)
    c_rate.add_data(d_rate, titles_from_data=True)
    c_rate.set_categories(cats)
    c_rate.series[0].graphicalProperties.line.solidFill = 'C00000'
    c_rate.series[0].graphicalProperties.line.width = 22000
    c_rate.series[0].smooth = True
    # Marker en cada punto
    c_rate.series[0].marker.symbol = 'circle'
    c_rate.series[0].marker.size = 5
    c_rate.series[0].marker.graphicalProperties.solidFill = 'C00000'
    c_rate.series[0].marker.graphicalProperties.line.solidFill = 'C00000'
    c_rate.series[0].dLbls = _line_labels('0.00%')
    ws.add_chart(c_rate, f'A{chart_start_row}')
    chart_start_row += 14

    # ── CHART 4: Abandonos diarios (barras naranjas) ──────────────────────
    c_aban = BarChart()
    c_aban.type = 'col'
    c_aban.grouping = 'clustered'
    c_aban.title = 'Abandonos por Día' if is_es else 'Daily Abandonments'
    c_aban.style = 2
    c_aban.height = 10; c_aban.width = 24
    c_aban.y_axis.title = 'Abandonos' if is_es else 'Abandonments'
    c_aban.y_axis.numFmt = '#,##0'
    c_aban.y_axis.crossAx = 100
    c_aban.x_axis.tickLblSkip = max(1, len(daily)//15)
    c_aban.x_axis.tickMarkSkip = max(1, len(daily)//15)
    c_aban.legend = None
    d_aban = Reference(ws, min_col=5, min_row=4, max_row=tr-1)
    c_aban.add_data(d_aban, titles_from_data=True)
    c_aban.set_categories(cats)
    c_aban.series[0].graphicalProperties.solidFill = 'BF8F00'
    c_aban.series[0].graphicalProperties.line.solidFill = '8C6800'
    c_aban.series[0].graphicalProperties.line.width = 6000
    c_aban.series[0].dLbls = _bar_labels('#,##0')
    ws.add_chart(c_aban, f'A{chart_start_row}')

# ─── Sheet 3: Gateway Analysis with chart ─────────────────────────────────────
def build_sheet3(wb, data):
    T = get_t(data)
    ws = wb.create_sheet(T['sheet3'])
    ws.sheet_view.showGridLines = False

    gateways   = data['gateways']
    operator   = data['operator']
    period     = f"{data['periodStart']} – {data['periodEnd']}"
    total_fail = data['totalFailures']

    ws.merge_cells('A1:G1')
    set_cell(ws, 1, 1, T['gw_title'], bg=BD, fg='FFFFFF', bold=True, sz=13, halign='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:G2')
    set_cell(ws, 2, 1, f'{operator}  |  Period: {period}', bg=BL, fg='555555', sz=9, halign='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    hdrs = T['gw_hdrs']
    col_widths = [20,14,14,14,16,10,46]
    for ci, (h, w) in enumerate(zip(hdrs, col_widths), 1):
        hdr(ws, 4, ci, h, bg=BM, wrap=True, width=w)
    ws.row_dimensions[4].height = 28

    total_inc = sum(g['total'] for g in gateways)

    def gw_diag(g):
        if total_fail > 0 and g['failures'] / total_fail > 0.8:
            return T['diag_dominant']
        if g['pending'] > 3: return T['diag_pend_many']
        if g['pending'] > 0: return T['diag_pend_one']
        if g['abandonments'] > 800: return T['diag_aban']
        return T['diag_monitor']

    for ri, g in enumerate(gateways, 5):
        is_top = ri == 5 and g['failures'] > 0
        bg = RL if is_top else (GR if ri % 2 == 0 else WH)
        fg = RD if is_top else '000000'
        vals = [g['gateway'], g['failures'], g['pending'], g['abandonments'],
                g['total'], g['pct']/100, gw_diag(g)]
        fmts = [None,'#,##0','#,##0','#,##0','#,##0','0.0%',None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            set_cell(ws, ri, ci, val, bg=bg, fg=fg, bold=is_top,
                     halign='left' if ci in [1,7] else 'center',
                     num_fmt=fmt, wrap=(ci==7))
        ws.row_dimensions[ri].height = 20 if is_top else 18

    # Total row
    tr = len(gateways) + 5
    tot_vals = ['TOTAL',
                sum(g['failures'] for g in gateways), sum(g['pending'] for g in gateways),
                sum(g['abandonments'] for g in gateways), total_inc, 1.0, '']
    for ci, (val, fmt) in enumerate(zip(tot_vals, [None,'#,##0','#,##0','#,##0','#,##0','0.0%',None]), 1):
        set_cell(ws, tr, ci, val, bg=BD, fg='FFFFFF', bold=True,
                 halign='left' if ci==1 else 'center', num_fmt=fmt)
    ws.row_dimensions[tr].height = 20

    # Chart: horizontal bar
    chart_r = tr + 2
    chart3 = BarChart()
    chart3.type = 'bar'
    chart3.title = T['gw_chart']
    chart3.style = 10; chart3.height = 14; chart3.width = 20
    d3 = Reference(ws, min_col=5, min_row=4, max_row=tr-1)
    cats3 = Reference(ws, min_col=1, min_row=5, max_row=tr-1)
    chart3.add_data(d3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.series[0].graphicalProperties.solidFill = '2E75B6'
    ws.add_chart(chart3, f'A{chart_r}')

# ─── Sheet 4: Channel & Device ─────────────────────────────────────────────────
def build_sheet4(wb, data):
    T = get_t(data)
    ws = wb.create_sheet(T['sheet4'])
    ws.sheet_view.showGridLines = False

    channels  = data['channels']
    platforms = data['platforms']
    operator  = data['operator']
    period    = f"{data['periodStart']} – {data['periodEnd']}"

    ws.merge_cells('A1:F1')
    set_cell(ws, 1, 1, T['ch_title'], bg=BD, fg='FFFFFF', bold=True, sz=13, halign='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:F2')
    set_cell(ws, 2, 1, f'{operator}  |  Period: {period}', bg=BL, fg='555555', sz=9, halign='center')
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 10

    ws.merge_cells('A4:F4')
    set_cell(ws, 4, 1, T['ch_by_channel'], bg=BM, fg='FFFFFF', bold=True, sz=10, halign='left')
    ws.row_dimensions[4].height = 22

    col_widths = [22,12,14,14,16,34]
    ch_hdrs = T['ch_hdrs']
    for ci, (h, w) in enumerate(zip(ch_hdrs, col_widths), 1):
        hdr(ws, 5, ci, h, bg=BM, width=w)
    ws.row_dimensions[5].height = 20

    tot_f = sum(c['failures'] for c in channels) or 1
    tot_a = sum(c['abandonments'] for c in channels) or 1

    for ri, ch in enumerate(channels, 6):
        is_top = ri == 6
        bg = AL if is_top else (GR if ri % 2 == 0 else WH)
        fg = AD if is_top else '000000'
        note = T['ch_note_top'] if is_top else T['ch_note_mon']
        vals = [ch['channel'], ch['failures'], ch['failures']/tot_f,
                ch['abandonments'], ch['abandonments']/tot_a, note]
        fmts = [None,'#,##0','0.0%','#,##0','0.0%',None]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 1):
            set_cell(ws, ri, ci, val, bg=bg, fg=fg, bold=is_top,
                     halign='left' if ci in [1,6] else 'center', num_fmt=fmt)
        ws.row_dimensions[ri].height = 18

    tr = len(channels) + 6
    for ci, (val, fmt) in enumerate(zip(['TOTAL',tot_f,1.0,tot_a,1.0,''],
                                         [None,'#,##0','0.0%','#,##0','0.0%',None]), 1):
        set_cell(ws, tr, ci, val, bg=BD, fg='FFFFFF', bold=True,
                 halign='left' if ci==1 else 'center', num_fmt=fmt)
    ws.row_dimensions[tr].height = 20

    # Platform section
    pl_start = tr + 2
    ws.merge_cells(f'A{pl_start}:F{pl_start}')
    set_cell(ws, pl_start, 1, T['pl_title'], bg=BM, fg='FFFFFF', bold=True, sz=10, halign='left')
    ws.row_dimensions[pl_start].height = 22

    pl_hdrs = T['pl_hdrs']
    for ci, h in enumerate(pl_hdrs, 1):
        hdr(ws, pl_start+1, ci, h, bg=BM)
    ws.row_dimensions[pl_start+1].height = 20

    for ri, p in enumerate(platforms, pl_start+2):
        bg = GR if ri % 2 == 0 else WH
        vals = [p['platform'], p['failures'], p['failures']/tot_f,
                p['abandonments'], p['abandonments']/tot_a, '']
        for ci, (val, fmt) in enumerate(zip(vals, [None,'#,##0','0.0%','#,##0','0.0%',None]), 1):
            set_cell(ws, ri, ci, val, bg=bg, halign='left' if ci==1 else 'center', num_fmt=fmt)
        ws.row_dimensions[ri].height = 18

# ─── Raw data sheets ──────────────────────────────────────────────────────────
def build_raw_sheet(wb, title, header_color, records, data, extra_cols=None):
    ws = wb.create_sheet(title[:31])
    ws.sheet_view.showGridLines = False

    T = get_t(data)
    cols   = list(T['raw_cols'])
    widths = [24,14,16,16,10,20,14,12,20,16,16]
    if extra_cols:
        cols   += [ec[0] for ec in extra_cols]
        widths += [ec[1] for ec in extra_cols]

    # Title row (styled)
    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    set_cell(ws, 1, 1, title, bg=header_color, fg='FFFFFF', bold=True, sz=11, halign='center')
    ws.row_dimensions[1].height = 28

    # Header row (styled)
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        hdr(ws, 2, ci, h, bg=header_color, width=w)
    ws.row_dimensions[2].height = 20

    # Data rows — use append() for speed (no per-cell styling on bulk rows)
    for r in records:
        row = [
            r.get('_pb',''), r.get('_date',''), r.get('_origin',''), r.get('_dest',''),
            r.get('_seat',''), r.get('_operator',''), r.get('_channel',''),
            r.get('_gateway',''), r.get('_pgStatus',''), r.get('_platform',''),
            r.get('_price', 0) or 0,
        ]
        if extra_cols:
            row += [r.get(ec[2],'') for ec in extra_cols]
        ws.append(row)

    # Column widths
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Autofilter + freeze
    ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}{2+len(records)}'
    ws.freeze_panes = 'B3'

# ─── Pending-failures sheet (special layout) ──────────────────────────────────
def build_pending_sheet(wb, records, data):
    T = get_t(data)
    ws = wb.create_sheet(T['sheet_pend'])
    ws.sheet_view.showGridLines = False

    cols   = T['pend_cols']
    widths = [24,14,16,16,10,20,14,12,16,34]

    ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
    set_cell(ws, 1, 1, T['pend_title'],
             bg=RD, fg='FFFFFF', bold=True, sz=11, halign='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f'A2:{get_column_letter(len(cols))}2')
    set_cell(ws, 2, 1,
             T['pend_warn'],
             bg=RL, fg=RD, bold=True, sz=10, halign='left', wrap=True)
    ws.row_dimensions[2].height = 24

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        hdr(ws, 3, ci, h, bg=RD, width=w)
    ws.row_dimensions[3].height = 20

    for r in records:
        ws.append([
            r.get('_pb',''), r.get('_date',''), r.get('_origin',''), r.get('_dest',''),
            r.get('_seat',''), r.get('_operator',''), r.get('_channel',''),
            r.get('_gateway',''), r.get('_price', 0) or 0, r.get('_email',''),
        ])

    ws.freeze_panes = 'B4'

# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    try:
        raw = sys.stdin.buffer.read()
        data = json.loads(raw.decode('utf-8'))

        # 'data' now IS the full payload from Node — pass it directly to build functions
        # Make sure lang is set
        if 'lang' not in data:
            data['lang'] = 'en'

        wb = Workbook()
        build_sheet1(wb, data)
        build_sheet2(wb, data)
        build_sheet3(wb, data)
        build_sheet4(wb, data)
        build_raw_sheet(wb,
            get_t(data)['raw_fail_title'](len(data.get('rawFailures', []))),
            RD, data.get('rawFailures', []), data)
        build_raw_sheet(wb,
            get_t(data)['raw_aban_title'](len(data.get('rawAbandon', []))),
            AD, data.get('rawAbandon', []), data)
        build_pending_sheet(wb, data.get('rawPending', []), data)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        sys.stdout.buffer.write(base64.b64encode(buf.read()))

    except Exception as e:
        import traceback
        sys.stderr.write(f"Error in main: {e}\n{traceback.format_exc()}\n")
        sys.exit(1)

if __name__ == '__main__':
    main()